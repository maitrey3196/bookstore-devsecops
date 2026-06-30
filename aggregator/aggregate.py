#!/usr/bin/env python3
"""
DevSecOps Security Report Aggregator - Version 2.3.2
Refined timing, Excel status, and cleanup.
Supports: Trivy FS, Trivy Image, Trivy Config, Semgrep
Ready for: OWASP ZAP, TruffleHog, Dependency Scanning
"""

import json
import sys
import logging
import time
from pathlib import Path
from datetime import datetime, timezone
from typing import Dict, Any, List, Optional
from dataclasses import dataclass, field
from enum import Enum

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, PageBreak, Image as PDFImage
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4

# --------------------------------------------------
# CONSTANTS
# --------------------------------------------------

SEVERITY_LEVELS = ["CRITICAL", "HIGH", "MEDIUM", "LOW"]

SEVERITY_COLORS = {
    "CRITICAL": "#CC0000",
    "HIGH": "#FF6600",
    "MEDIUM": "#FFCC00",
    "LOW": "#3399FF",
    "SEMGREP": "#9933CC"
}

# --------------------------------------------------
# ENUMS
# --------------------------------------------------

class ScannerStatus(Enum):
    SUCCESS = "SUCCESS"
    FAILED = "FAILED"
    NOT_FOUND = "NOT_FOUND"
    NOT_RUN = "NOT_RUN"

# --------------------------------------------------
# CONFIGURATION
# --------------------------------------------------

@dataclass
class Config:
    report_dir: Path = Path("/var/jenkins_home/reports/raw")
    final_dir: Path = Path("/var/jenkins_home/reports/final")
    report_files: Dict[str, Path] = field(default_factory=dict)
    output_files: Dict[str, Path] = field(default_factory=dict)
    fail_on_critical: bool = True
    fail_on_high: bool = False
    fail_on_any_finding: bool = False
    max_critical_allowed: int = 0
    max_high_allowed: int = 10
    log_level: str = "INFO"
    log_file: Optional[Path] = None

    def __post_init__(self):
        if not self.report_files:
            self.report_files = {
                "trivy_fs": self.report_dir / "trivy-fs.json",
                "trivy_image": self.report_dir / "trivy-image.json",
                "trivy_config": self.report_dir / "trivy-config.json",
                "semgrep": self.report_dir / "semgrep.json"
            }
        if not self.output_files:
            self.output_files = {
                "json": self.final_dir / "security-summary.json",
                "excel": self.final_dir / "security-report.xlsx",
                "pdf": self.final_dir / "security-report.pdf",
                "html": self.final_dir / "security-report.html",
                "chart_security": self.final_dir / "security-chart.png",
                "chart_config": self.final_dir / "config-chart.png"
            }
        if not self.log_file:
            self.log_file = self.final_dir / "aggregator.log"

# --------------------------------------------------
# DATA MODELS
# --------------------------------------------------

@dataclass
class ScannerResult:
    name: str
    status: str = ScannerStatus.NOT_RUN.value
    findings: Dict[str, int] = field(default_factory=dict)
    total: int = 0
    execution_time: float = 0.0
    scan_date: str = ""
    report_size: int = 0
    errors: List[str] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)

    def __post_init__(self):
        if not self.scan_date:
            self.scan_date = datetime.now(timezone.utc).isoformat()

@dataclass
class SecuritySummary:
    application: Dict[str, int] = field(default_factory=lambda: {"CRITICAL":0,"HIGH":0,"MEDIUM":0,"LOW":0})
    configuration: Dict[str, int] = field(default_factory=lambda: {"CRITICAL":0,"HIGH":0,"MEDIUM":0,"LOW":0})
    sast: Dict[str, int] = field(default_factory=lambda: {"findings":0})

    def get_total_app_vulns(self) -> int: return sum(self.application.values())
    def get_total_config_vulns(self) -> int: return sum(self.configuration.values())
    def get_total_critical(self) -> int: return self.application["CRITICAL"] + self.configuration["CRITICAL"]
    def get_total_high(self) -> int: return self.application["HIGH"] + self.configuration["HIGH"]
    def get_grand_total(self) -> int: return self.get_total_app_vulns() + self.get_total_config_vulns() + self.sast["findings"]

# --------------------------------------------------
# LOGGING
# --------------------------------------------------

def setup_logging(config: Config) -> logging.Logger:
    logger = logging.getLogger("DevSecOpsAggregator")
    logger.setLevel(getattr(logging, config.log_level))
    if logger.handlers:
        return logger

    console = logging.StreamHandler()
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', '%H:%M:%S'))
    logger.addHandler(console)

    if config.log_file:
        config.log_file.parent.mkdir(parents=True, exist_ok=True)
        file_h = logging.FileHandler(config.log_file)
        file_h.setLevel(logging.DEBUG)
        file_h.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s', '%Y-%m-%d %H:%M:%S'))
        logger.addHandler(file_h)
    return logger

# --------------------------------------------------
# PARSERS (unchanged)
# --------------------------------------------------

def count_vulnerabilities(vulns: List[Dict]) -> Dict[str, int]:
    c = {s:0 for s in SEVERITY_LEVELS}
    for v in vulns:
        sev = v.get("Severity","").upper()
        if sev in c: c[sev] += 1
    return c

def count_misconfigurations(misconfigs: List[Dict]) -> Dict[str, int]:
    c = {s:0 for s in SEVERITY_LEVELS}
    for m in misconfigs:
        sev = m.get("Severity","").upper()
        if sev in c: c[sev] += 1
    return c

def detect_trivy_format(data: Dict) -> str:
    if "Results" in data:
        if isinstance(data["Results"], list) and data["Results"]:
            if "Misconfigurations" in data["Results"][0]: return "v1_results_with_misconfigs"
            if "Vulnerabilities" in data["Results"][0]: return "v1_results_with_vulns"
        return "v1_results"
    if "Findings" in data: return "v2_findings"
    if "Misconfigurations" in data: return "v3_misconfigs"
    return "unknown"

def extract_findings_generic(data: Dict, logger: logging.Logger) -> Dict[str, int]:
    logger.warning("Generic extraction – may overcount if summary/detail both contain severity")
    findings = {s:0 for s in SEVERITY_LEVELS}
    def rec(obj, d=0):
        if d>10: return
        if isinstance(obj, dict):
            if "Severity" in obj:
                sev = obj["Severity"].upper()
                if sev in findings: findings[sev] += 1
            for v in obj.values(): rec(v, d+1)
        elif isinstance(obj, list):
            for i in obj: rec(i, d+1)
    rec(data)
    return findings

def parse_trivy_fs(file_path: Path, logger: logging.Logger) -> ScannerResult:
    scanner = ScannerResult(name="Trivy Filesystem", status=ScannerStatus.NOT_RUN.value)
    start = time.time()
    try:
        if not file_path.exists():
            scanner.status = ScannerStatus.NOT_FOUND.value
            scanner.errors.append(f"File not found: {file_path}")
            logger.warning(f"Trivy FS report not found at {file_path}")
            return scanner
        with open(file_path) as f:
            data = json.load(f)
        scanner.report_size = file_path.stat().st_size
        findings = {s:0 for s in SEVERITY_LEVELS}
        fmt = detect_trivy_format(data)
        logger.debug(f"Trivy FS format: {fmt}")
        if fmt in ("v1_results_with_vulns","v1_results"):
            for res in data.get("Results",[]):
                vulns = count_vulnerabilities(res.get("Vulnerabilities",[]))
                for s in SEVERITY_LEVELS: findings[s] += vulns[s]
        else:
            logger.warning("Unknown Trivy FS format, generic extraction")
            findings = extract_findings_generic(data, logger)
        scanner.findings = findings
        scanner.total = sum(findings.values())
        scanner.status = ScannerStatus.SUCCESS.value
        scanner.execution_time = time.time() - start
        logger.info(f"✓ Trivy FS: {scanner.total} findings ({scanner.execution_time:.2f}s)")
    except json.JSONDecodeError as e:
        scanner.status = ScannerStatus.FAILED.value
        scanner.errors.append(f"Invalid JSON: {str(e)}")
        logger.error(f"✗ Trivy FS: Invalid JSON - {str(e)}")
    except Exception as e:
        scanner.status = ScannerStatus.FAILED.value
        scanner.errors.append(str(e))
        logger.error(f"✗ Trivy FS: {str(e)}")
    return scanner

def parse_trivy_image(file_path: Path, logger: logging.Logger) -> ScannerResult:
    scanner = ScannerResult(name="Trivy Image", status=ScannerStatus.NOT_RUN.value)
    start = time.time()
    try:
        if not file_path.exists():
            scanner.status = ScannerStatus.NOT_FOUND.value
            scanner.errors.append(f"File not found: {file_path}")
            logger.warning(f"Trivy Image report not found at {file_path}")
            return scanner
        with open(file_path) as f:
            data = json.load(f)
        scanner.report_size = file_path.stat().st_size
        findings = {s:0 for s in SEVERITY_LEVELS}
        fmt = detect_trivy_format(data)
        logger.debug(f"Trivy Image format: {fmt}")
        if fmt in ("v1_results_with_vulns","v1_results"):
            for res in data.get("Results",[]):
                vulns = count_vulnerabilities(res.get("Vulnerabilities",[]))
                for s in SEVERITY_LEVELS: findings[s] += vulns[s]
        else:
            logger.warning("Unknown Trivy Image format, generic extraction")
            findings = extract_findings_generic(data, logger)
        scanner.findings = findings
        scanner.total = sum(findings.values())
        scanner.status = ScannerStatus.SUCCESS.value
        scanner.execution_time = time.time() - start
        logger.info(f"✓ Trivy Image: {scanner.total} findings ({scanner.execution_time:.2f}s)")
    except json.JSONDecodeError as e:
        scanner.status = ScannerStatus.FAILED.value
        scanner.errors.append(f"Invalid JSON: {str(e)}")
        logger.error(f"✗ Trivy Image: Invalid JSON - {str(e)}")
    except Exception as e:
        scanner.status = ScannerStatus.FAILED.value
        scanner.errors.append(str(e))
        logger.error(f"✗ Trivy Image: {str(e)}")
    return scanner

def parse_trivy_config(file_path: Path, logger: logging.Logger) -> ScannerResult:
    scanner = ScannerResult(name="Trivy Config", status=ScannerStatus.NOT_RUN.value)
    start = time.time()
    try:
        if not file_path.exists():
            scanner.status = ScannerStatus.NOT_FOUND.value
            scanner.errors.append(f"File not found: {file_path}")
            logger.warning(f"Trivy Config report not found at {file_path}")
            return scanner
        with open(file_path) as f:
            data = json.load(f)
        scanner.report_size = file_path.stat().st_size
        findings = {s:0 for s in SEVERITY_LEVELS}
        fmt = detect_trivy_format(data)
        logger.info(f"Trivy Config format: {fmt}")
        if fmt == "v1_results_with_misconfigs":
            for res in data.get("Results",[]):
                mis = count_misconfigurations(res.get("Misconfigurations",[]))
                for s in SEVERITY_LEVELS: findings[s] += mis[s]
        elif fmt == "v3_misconfigs":
            findings = count_misconfigurations(data.get("Misconfigurations",[]))
        elif fmt == "v2_findings":
            findings = count_misconfigurations(data.get("Findings",[]))
        elif fmt == "v1_results":
            for res in data.get("Results",[]):
                lst = res.get("Misconfigurations",[]) or res.get("Findings",[])
                if lst:
                    mis = count_misconfigurations(lst)
                    for s in SEVERITY_LEVELS: findings[s] += mis[s]
        else:
            logger.warning("Unknown Trivy Config structure, generic extraction (may overcount)")
            findings = extract_findings_generic(data, logger)
        scanner.findings = findings
        scanner.total = sum(findings.values())
        scanner.status = ScannerStatus.SUCCESS.value
        scanner.execution_time = time.time() - start
        logger.info(f"✓ Trivy Config: {scanner.total} findings ({scanner.execution_time:.2f}s)")
    except json.JSONDecodeError as e:
        scanner.status = ScannerStatus.FAILED.value
        scanner.errors.append(f"Invalid JSON: {str(e)}")
        logger.error(f"✗ Trivy Config: Invalid JSON - {str(e)}")
    except Exception as e:
        scanner.status = ScannerStatus.FAILED.value
        scanner.errors.append(str(e))
        logger.error(f"✗ Trivy Config: {str(e)}")
    return scanner

def parse_semgrep(file_path: Path, logger: logging.Logger) -> ScannerResult:
    scanner = ScannerResult(name="Semgrep SAST", status=ScannerStatus.NOT_RUN.value)
    start = time.time()
    try:
        if not file_path.exists():
            scanner.status = ScannerStatus.NOT_FOUND.value
            scanner.errors.append(f"File not found: {file_path}")
            logger.warning(f"Semgrep report not found at {file_path}")
            return scanner
        with open(file_path) as f:
            data = json.load(f)
        scanner.report_size = file_path.stat().st_size
        findings_count = len(data.get("results",[]))
        scanner.findings = {"findings": findings_count}
        scanner.total = findings_count
        scanner.status = ScannerStatus.SUCCESS.value
        scanner.execution_time = time.time() - start
        if "version" in data:
            scanner.metadata["version"] = data["version"]
        logger.info(f"✓ Semgrep: {findings_count} findings ({scanner.execution_time:.2f}s)")
    except json.JSONDecodeError as e:
        scanner.status = ScannerStatus.FAILED.value
        scanner.errors.append(f"Invalid JSON: {str(e)}")
        logger.error(f"✗ Semgrep: Invalid JSON - {str(e)}")
    except Exception as e:
        scanner.status = ScannerStatus.FAILED.value
        scanner.errors.append(str(e))
        logger.error(f"✗ Semgrep: {str(e)}")
    return scanner

# --------------------------------------------------
# CHART GENERATION
# --------------------------------------------------

def generate_charts(summary: SecuritySummary, config: Config, logger: logging.Logger) -> List[str]:
    logger.info("Generating charts...")
    charts = []
    try:
        if summary.get_total_app_vulns() > 0 or summary.sast["findings"] > 0:
            fig, ax = plt.subplots(figsize=(10,6))
            labels = list(summary.application.keys()) + ["Semgrep"]
            values = list(summary.application.values()) + [summary.sast["findings"]]
            colors_list = [SEVERITY_COLORS.get(s,"#999") for s in summary.application.keys()] + [SEVERITY_COLORS["SEMGREP"]]
            bars = ax.bar(labels, values, color=colors_list, edgecolor='black', linewidth=0.5)
            for bar, val in zip(bars, values):
                if val>0: ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.5, str(val), ha='center', va='bottom', fontweight='bold')
            ax.set_title('Application Security Vulnerabilities', fontsize=16, fontweight='bold', pad=20)
            ax.set_xlabel('Severity Level', fontsize=12, fontweight='bold')
            ax.set_ylabel('Number of Findings', fontsize=12, fontweight='bold')
            ax.grid(axis='y', alpha=0.3, linestyle='--')
            ax.set_facecolor('#F8F9FA')
            fig.patch.set_facecolor('white')
            plt.tight_layout()
            plt.savefig(config.output_files["chart_security"], dpi=150, bbox_inches='tight')
            plt.close()
            charts.append("security-chart.png")
            logger.info("✓ Security chart generated")
        if summary.get_total_config_vulns() > 0:
            fig, ax = plt.subplots(figsize=(10,6))
            labels = [f"Config\n{s}" for s in summary.configuration.keys()]
            values = list(summary.configuration.values())
            colors_list = [SEVERITY_COLORS.get(s,"#999") for s in summary.configuration.keys()]
            bars = ax.bar(labels, values, color=colors_list, edgecolor='black', linewidth=0.5)
            for bar, val in zip(bars, values):
                if val>0: ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.5, str(val), ha='center', va='bottom', fontweight='bold')
            ax.set_title('Infrastructure Misconfigurations', fontsize=16, fontweight='bold', pad=20)
            ax.set_xlabel('Severity Level', fontsize=12, fontweight='bold')
            ax.set_ylabel('Number of Misconfigurations', fontsize=12, fontweight='bold')
            ax.grid(axis='y', alpha=0.3, linestyle='--')
            ax.set_facecolor('#F8F9FA')
            fig.patch.set_facecolor('white')
            plt.tight_layout()
            plt.savefig(config.output_files["chart_config"], dpi=150, bbox_inches='tight')
            plt.close()
            charts.append("config-chart.png")
            logger.info("✓ Config chart generated")
    except Exception as e:
        logger.error(f"✗ Chart generation failed: {e}", exc_info=True)
    return charts

# --------------------------------------------------
# REPORT GENERATORS
# --------------------------------------------------

def generate_json_report(summary, scanners, config, logger, parsing_duration, report_gen_duration, total_runtime):
    logger.info("Generating JSON report...")
    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "generator_version": "2.3.2",
        "timing": {
            "parsing_duration_seconds": round(parsing_duration, 2),
            "report_generation_seconds": round(report_gen_duration, 2),
            "total_runtime_seconds": round(total_runtime, 2)
        },
        "summary": {
            "application_vulnerabilities": summary.application,
            "configuration_misconfigurations": summary.configuration,
            "sast_findings": summary.sast,
            "totals": {
                "application": summary.get_total_app_vulns(),
                "configuration": summary.get_total_config_vulns(),
                "sast": summary.sast["findings"],
                "critical_total": summary.get_total_critical(),
                "high_total": summary.get_total_high(),
                "grand_total": summary.get_grand_total()
            }
        },
        "scanners": {}
    }
    for name, sc in scanners.items():
        report["scanners"][name] = {
            "status": sc.status,
            "findings": sc.findings,
            "total": sc.total,
            "execution_time": round(sc.execution_time,3),
            "scan_date": sc.scan_date,
            "report_size_bytes": sc.report_size,
            "errors": sc.errors,
            "metadata": sc.metadata
        }
    with open(config.output_files["json"], "w") as f:
        json.dump(report, f, indent=4)
    logger.info(f"✓ JSON report generated: {config.output_files['json']}")

def generate_excel_report(summary, scanners, config, logger):
    logger.info("Generating Excel report...")
    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    section_font = Font(name='Arial', size=11, bold=True, color='2F5496')
    normal_font = Font(name='Arial', size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')

    def severity_status(metric, count):
        if count == 0:
            return "✅ OK"
        if "CRITICAL" in metric:
            return "⚠ Immediate Action Required"
        if "HIGH" in metric:
            return "⚠ Action Required"
        if "MEDIUM" in metric:
            return "ℹ Review Needed"
        if "LOW" in metric:
            return "ℹ Monitor"
        if "SEMGREP" in metric:
            return "⚠ Code Issues Found" if count > 0 else "✅ Code Clean"
        return ""

    # Sheet 1 Summary
    ws = wb.create_sheet("Summary")
    ws.merge_cells('A1:C1')
    ws['A1'].value = "DevSecOps Security Report"
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color='1F3864')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A3'] = "Generated At:"; ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['A4'] = "Generator Version:"; ws['B4'] = "2.3.2"
    row = 6
    # App section
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].value = "Application Vulnerabilities"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    row += 1
    for col, h in enumerate(['Metric','Count','Status'],1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = border
    app_data = [("CRITICAL", summary.application["CRITICAL"]), ("HIGH", summary.application["HIGH"]),
                ("MEDIUM", summary.application["MEDIUM"]), ("LOW", summary.application["LOW"]),
                ("SEMGREP FINDINGS", summary.sast["findings"])]
    for i, (metric, count) in enumerate(app_data, start=row+1):
        ws.cell(row=i, column=1, value=metric).font = normal_font
        c = ws.cell(row=i, column=2, value=count); c.font = normal_font; c.alignment = center_align
        ws.cell(row=i, column=3, value=severity_status(metric, count)).font = normal_font
        for col in range(1,4): ws.cell(row=i, column=col).border = border
        if "CRITICAL" in metric and count>0:
            ws.cell(row=i, column=1).font = Font(name='Arial', size=10, bold=True, color='CC0000')
            c.font = Font(name='Arial', size=10, bold=True, color='CC0000')
        elif "HIGH" in metric and count>0:
            ws.cell(row=i, column=1).font = Font(name='Arial', size=10, bold=True, color='FF6600')
            c.font = Font(name='Arial', size=10, bold=True, color='FF6600')
    row = row + len(app_data) + 2
    # Config section
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].value = "Infrastructure Misconfigurations"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    row += 1
    for col, h in enumerate(['Metric','Count','Status'],1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = border
    config_data = [("CONFIG CRITICAL", summary.configuration["CRITICAL"]), ("CONFIG HIGH", summary.configuration["HIGH"]),
                   ("CONFIG MEDIUM", summary.configuration["MEDIUM"]), ("CONFIG LOW", summary.configuration["LOW"])]
    for i, (metric, count) in enumerate(config_data, start=row+1):
        ws.cell(row=i, column=1, value=metric).font = normal_font
        c = ws.cell(row=i, column=2, value=count); c.font = normal_font; c.alignment = center_align
        ws.cell(row=i, column=3, value=severity_status(metric, count)).font = normal_font
        for col in range(1,4): ws.cell(row=i, column=col).border = border
        if "CRITICAL" in metric and count>0:
            ws.cell(row=i, column=1).font = Font(name='Arial', size=10, bold=True, color='CC0000')
            c.font = Font(name='Arial', size=10, bold=True, color='CC0000')
        elif "HIGH" in metric and count>0:
            ws.cell(row=i, column=1).font = Font(name='Arial', size=10, bold=True, color='FF6600')
            c.font = Font(name='Arial', size=10, bold=True, color='FF6600')
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30

    # Sheet 2 Scanner Status
    ws2 = wb.create_sheet("Scanner Status")
    ws2['A1'] = "Scanner Status Report"; ws2['A1'].font = Font(name='Arial', size=14, bold=True)
    headers = ["Scanner","Status","Total Findings","Execution Time","Report Size","Errors"]
    for col, h in enumerate(headers,1):
        cell = ws2.cell(row=3, column=col, value=h); cell.font = header_font; cell.fill = header_fill; cell.border = border
    for i, (name, sc) in enumerate(scanners.items(), start=4):
        ws2.cell(row=i, column=1, value=sc.name).font = normal_font
        scell = ws2.cell(row=i, column=2, value=sc.status); scell.font = normal_font
        ws2.cell(row=i, column=3, value=sc.total).font = normal_font
        ws2.cell(row=i, column=4, value=f"{sc.execution_time:.2f}s").font = normal_font
        ws2.cell(row=i, column=5, value=f"{sc.report_size} bytes").font = normal_font
        ws2.cell(row=i, column=6, value=", ".join(sc.errors) if sc.errors else "None").font = normal_font
        for col in range(1,7): ws2.cell(row=i, column=col).border = border
        if sc.status == ScannerStatus.SUCCESS.value:
            scell.font = Font(name='Arial', size=10, color='006600', bold=True)
        elif sc.status in (ScannerStatus.FAILED.value, ScannerStatus.NOT_FOUND.value):
            scell.font = Font(name='Arial', size=10, color='CC0000', bold=True)
    for col in range(1,7): ws2.column_dimensions[get_column_letter(col)].width = 20

    # Sheet 3 Errors
    ws3 = wb.create_sheet("Errors")
    ws3['A1'] = "Error Report"; ws3['A1'].font = Font(name='Arial', size=14, bold=True)
    erow = 3; has_err = False
    for sc in scanners.values():
        if sc.errors:
            has_err = True
            ws3.cell(row=erow, column=1, value=sc.name).font = Font(name='Arial', size=10, bold=True)
            erow+=1
            for e in sc.errors:
                ws3.cell(row=erow, column=1, value="").font = normal_font
                ws3.cell(row=erow, column=2, value=e).font = normal_font
                erow+=1
            erow+=1
    if not has_err: ws3.cell(row=3, column=1, value="No errors reported").font = normal_font
    ws3.column_dimensions['A'].width = 25; ws3.column_dimensions['B'].width = 80

    wb.save(config.output_files["excel"])
    logger.info(f"✓ Excel report generated: {config.output_files['excel']}")

def generate_pdf_report(summary, scanners, config, logger, charts_generated):
    logger.info("Generating PDF report...")
    try:
        doc = SimpleDocTemplate(str(config.output_files["pdf"]), pagesize=A4,
                                rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
        styles = getSampleStyleSheet()
        section_style = ParagraphStyle('SectionHeader', parent=styles['Heading2'],
                                       textColor=colors.HexColor('#2F5496'), spaceAfter=12, spaceBefore=12)
        elements = []
        # Policy banner
        total_critical = summary.get_total_critical()
        if total_critical > 0:
            elements.append(Paragraph(f"<font color='red'><b>⚠ POLICY FAILED: {total_critical} critical findings</b></font>", styles["Heading1"]))
        else:
            elements.append(Paragraph("<font color='green'><b>✓ POLICY PASSED: No critical findings</b></font>", styles["Heading1"]))
        elements.append(Spacer(1,20))
        # Executive summary
        elements.append(Paragraph("Executive Summary", section_style))
        elements.append(Paragraph(f"A total of <b>{summary.get_grand_total()}</b> security findings were identified.", styles["Normal"]))
        elements.append(Spacer(1,20))
        # Charts
        elements.append(Paragraph("Vulnerability Distribution", section_style))
        if "security-chart.png" in charts_generated:
            elements.append(PDFImage(str(config.output_files["chart_security"]), width=450, height=280))
            elements.append(Spacer(1,10))
        else:
            elements.append(Paragraph("<i>Application vulnerability chart not available.</i>", styles["Normal"]))
        if "config-chart.png" in charts_generated:
            elements.append(PDFImage(str(config.output_files["chart_config"]), width=450, height=280))
            elements.append(Spacer(1,10))
        else:
            elements.append(Paragraph("<i>Infrastructure misconfiguration chart not available.</i>", styles["Normal"]))
        elements.append(PageBreak())
        # App table
        elements.append(Paragraph("Application Vulnerabilities", section_style))
        app_table = Table([["Severity","Count"],
                           ["CRITICAL", str(summary.application["CRITICAL"])],
                           ["HIGH", str(summary.application["HIGH"])],
                           ["MEDIUM", str(summary.application["MEDIUM"])],
                           ["LOW", str(summary.application["LOW"])],
                           ["Semgrep Findings", str(summary.sast["findings"])]], colWidths=[200,100])
        app_table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor('#2F5496')),
                                       ("TEXTCOLOR",(0,0),(-1,0),colors.white),
                                       ("GRID",(0,0),(-1,-1),0.5,colors.grey),
                                       ("ALIGN",(0,0),(-1,-1),"CENTER")]))
        elements.append(app_table)
        elements.append(Spacer(1,20))
        # Config table
        elements.append(Paragraph("Infrastructure Misconfigurations", section_style))
        conf_table = Table([["Severity","Count"],
                            ["CONFIG CRITICAL", str(summary.configuration["CRITICAL"])],
                            ["CONFIG HIGH", str(summary.configuration["HIGH"])],
                            ["CONFIG MEDIUM", str(summary.configuration["MEDIUM"])],
                            ["CONFIG LOW", str(summary.configuration["LOW"])]], colWidths=[200,100])
        conf_table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor('#2F5496')),
                                        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
                                        ("GRID",(0,0),(-1,-1),0.5,colors.grey),
                                        ("ALIGN",(0,0),(-1,-1),"CENTER")]))
        elements.append(conf_table)
        elements.append(Spacer(1,20))
        # Recommendations
        elements.append(Paragraph("Recommendations", section_style))
        if summary.get_total_critical() > 0:
            elements.append(Paragraph("• Address all CRITICAL findings immediately", styles["Normal"]))
        if summary.get_total_high() > config.max_high_allowed:
            elements.append(Paragraph(f"• Prioritize HIGH findings (>{config.max_high_allowed}) in next sprint", styles["Normal"]))
        if summary.sast["findings"] > 20:  # could be configurable later
            elements.append(Paragraph("• Review high-priority code quality issues", styles["Normal"]))
        if not any(s.status == ScannerStatus.SUCCESS.value for s in scanners.values()):
            elements.append(Paragraph("• No scans completed successfully – review configurations", styles["Normal"]))
        doc.build(elements)
        logger.info(f"✓ PDF report generated: {config.output_files['pdf']}")
    except Exception as e:
        logger.error(f"✗ PDF generation failed: {e}")
        raise

def generate_html_report(summary, scanners, config, logger, charts_generated):
    logger.info("Generating HTML report...")
    total_crit = summary.get_total_critical()
    banner = f'<div style="background:#dc3545;color:white;padding:15px;text-align:center;margin-bottom:20px;border-radius:5px;"><strong>⚠ POLICY FAILED: {total_crit} critical findings</strong></div>' if total_crit>0 else '<div style="background:#28a745;color:white;padding:15px;text-align:center;margin-bottom:20px;border-radius:5px;"><strong>✓ POLICY PASSED: No critical findings</strong></div>'
    html = f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><title>DevSecOps Security Report</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Segoe UI',sans-serif;background:linear-gradient(135deg,#667eea,#764ba2);padding:20px;color:#333}}
.container{{max-width:1200px;margin:0 auto;background:white;border-radius:10px;box-shadow:0 10px 40px rgba(0,0,0,0.1);overflow:hidden}}
.header{{background:linear-gradient(135deg,#2F5496,#1a365d);color:white;padding:30px;text-align:center}}
.header h1{{font-size:2.5em;margin-bottom:10px}}
.content{{padding:30px}}
.section{{margin-bottom:40px}}
.section h2{{color:#2F5496;border-bottom:3px solid #2F5496;padding-bottom:10px;margin-bottom:20px}}
table{{width:100%;border-collapse:collapse;margin-bottom:20px;box-shadow:0 2px 10px rgba(0,0,0,0.05)}}
thead{{background:#2F5496;color:white}}
th{{padding:12px;text-align:left;font-weight:600;text-transform:uppercase;font-size:0.9em}}
td{{padding:12px;border-bottom:1px solid #e0e0e0}}
.critical{{color:#CC0000;font-weight:bold}}
.high{{color:#FF6600;font-weight:bold}}
.medium{{color:#FFCC00;font-weight:bold}}
.low{{color:#3399FF;font-weight:bold}}
.status-success{{color:#28a745;font-weight:bold}}
.status-error{{color:#dc3545;font-weight:bold}}
.info-box{{background:#f8f9fa;border-left:4px solid #2F5496;padding:15px;margin-bottom:20px;border-radius:4px}}
.chart-container{{text-align:center;margin:20px 0;padding:20px;background:#f8f9fa;border-radius:8px}}
.chart-container img{{max-width:100%;height:auto;border-radius:4px;box-shadow:0 4px 8px rgba(0,0,0,0.1)}}
.chart-unavailable{{text-align:center;padding:40px;background:#f8f9fa;border-radius:8px;color:#666;font-style:italic}}
.footer{{background:#f8f9fa;text-align:center;padding:20px;color:#666;border-top:1px solid #e0e0e0}}
</style></head>
<body><div class="container"><div class="header"><h1>🛡️ DevSecOps Security Report</h1><p>Comprehensive Security Analysis Summary</p></div><div class="content">
{banner}
<div class="info-box"><p><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p><p><strong>Generator Version:</strong> 2.3.2</p><p><strong>Scanners:</strong> {', '.join(s.name for s in scanners.values())}</p></div>
<div class="section"><h2>📊 Vulnerability Distribution</h2>"""
    if "security-chart.png" in charts_generated:
        html += f'<div class="chart-container"><h3>Application Vulnerabilities</h3><img src="{config.output_files["chart_security"].name}" alt="Security Chart"></div>'
    else:
        html += '<div class="chart-unavailable"><p>Application vulnerability chart not available.</p></div>'
    if "config-chart.png" in charts_generated:
        html += f'<div class="chart-container"><h3>Infrastructure Misconfigurations</h3><img src="{config.output_files["chart_config"].name}" alt="Config Chart"></div>'
    else:
        html += '<div class="chart-unavailable"><p>Infrastructure misconfiguration chart not available.</p></div>'
    html += f"""</div>
<div class="section"><h2>🔍 Application Vulnerabilities</h2><table><thead><tr><th>Severity</th><th>Count</th><th>Status</th></tr></thead><tbody>
<tr><td class="critical">🔴 CRITICAL</td><td>{summary.application['CRITICAL']}</td><td>{'⚠ Immediate Action Required' if summary.application['CRITICAL']>0 else '✅ No Issues'}</td></tr>
<tr><td class="high">🟠 HIGH</td><td>{summary.application['HIGH']}</td><td>{'⚠ Action Required' if summary.application['HIGH']>0 else '✅ No Issues'}</td></tr>
<tr><td class="medium">🟡 MEDIUM</td><td>{summary.application['MEDIUM']}</td><td>{'ℹ Review Needed' if summary.application['MEDIUM']>0 else '✅ No Issues'}</td></tr>
<tr><td class="low">🔵 LOW</td><td>{summary.application['LOW']}</td><td>{'ℹ Monitor' if summary.application['LOW']>0 else '✅ No Issues'}</td></tr>
<tr><td><strong>🟣 SEMGREP FINDINGS</strong></td><td>{summary.sast['findings']}</td><td>{'⚠ Code Issues Found' if summary.sast['findings']>0 else '✅ Code Clean'}</td></tr>
</tbody></table></div>
<div class="section"><h2>⚙️ Infrastructure Misconfigurations</h2><table><thead><tr><th>Severity</th><th>Count</th><th>Status</th></tr></thead><tbody>
<tr><td class="critical">🔴 CONFIG CRITICAL</td><td>{summary.configuration['CRITICAL']}</td><td>{'⚠ Immediate Action Required' if summary.configuration['CRITICAL']>0 else '✅ Compliant'}</td></tr>
<tr><td class="high">🟠 CONFIG HIGH</td><td>{summary.configuration['HIGH']}</td><td>{'⚠ Action Required' if summary.configuration['HIGH']>0 else '✅ Compliant'}</td></tr>
<tr><td class="medium">🟡 CONFIG MEDIUM</td><td>{summary.configuration['MEDIUM']}</td><td>{'ℹ Review Needed' if summary.configuration['MEDIUM']>0 else '✅ Compliant'}</td></tr>
<tr><td class="low">🔵 CONFIG LOW</td><td>{summary.configuration['LOW']}</td><td>{'ℹ Monitor' if summary.configuration['LOW']>0 else '✅ Compliant'}</td></tr>
</tbody></table></div>
<div class="section"><h2>📈 Scanner Status</h2><table><thead><tr><th>Scanner</th><th>Status</th><th>Findings</th><th>Time</th></tr></thead><tbody>"""
    for sc in scanners.values():
        cls = "status-success" if sc.status == ScannerStatus.SUCCESS.value else "status-error"
        html += f"<tr><td>{sc.name}</td><td class='{cls}'>{sc.status}</td><td>{sc.total}</td><td>{sc.execution_time:.2f}s</td></tr>"
    html += """</tbody></table></div></div><div class="footer"><p>🔒 DevSecOps Security Pipeline • Automated Report Generation</p><p>Generated by DevSecOps Aggregator v2.3.2</p></div></div></body></html>"""
    with open(config.output_files["html"], "w") as f:
        f.write(html)
    logger.info(f"✓ HTML report generated: {config.output_files['html']}")

# --------------------------------------------------
# MAIN EXECUTION (fixed timing order)
# --------------------------------------------------

def main() -> int:
    overall_start = time.time()
    config = Config()
    logger = setup_logging(config)
    logger.info("="*60)
    logger.info("🛡️  DEVSECOPS SECURITY REPORT AGGREGATOR v2.3.2")
    logger.info("="*60)

    try:
        config.final_dir.mkdir(parents=True, exist_ok=True)
        # Phase 1: Parsing
        logger.info("Phase 1: Parsing Security Reports")
        scanners = {
            "trivy_fs": parse_trivy_fs(config.report_files["trivy_fs"], logger),
            "trivy_image": parse_trivy_image(config.report_files["trivy_image"], logger),
            "trivy_config": parse_trivy_config(config.report_files["trivy_config"], logger),
            "semgrep": parse_semgrep(config.report_files["semgrep"], logger)
        }
        parsing_end = time.time()

        # Phase 2: Aggregate
        summary = SecuritySummary()
        if scanners["trivy_fs"].status == ScannerStatus.SUCCESS.value:
            for s in SEVERITY_LEVELS: summary.application[s] += scanners["trivy_fs"].findings.get(s,0)
        if scanners["trivy_image"].status == ScannerStatus.SUCCESS.value:
            for s in SEVERITY_LEVELS: summary.application[s] += scanners["trivy_image"].findings.get(s,0)
        if scanners["trivy_config"].status == ScannerStatus.SUCCESS.value:
            for s in SEVERITY_LEVELS: summary.configuration[s] += scanners["trivy_config"].findings.get(s,0)
        if scanners["semgrep"].status == ScannerStatus.SUCCESS.value:
            summary.sast["findings"] = scanners["semgrep"].findings.get("findings",0)

        # Phase 3: Generate reports – order: Excel, Charts, PDF, HTML, JSON (all included in timing)
        logger.info("Phase 3: Generating Reports")
        generate_excel_report(summary, scanners, config, logger)
        charts_generated = generate_charts(summary, config, logger)
        generate_pdf_report(summary, scanners, config, logger, charts_generated)
        generate_html_report(summary, scanners, config, logger, charts_generated)
        generate_json_report(summary, scanners, config, logger,
                             parsing_duration=parsing_end - overall_start,
                             report_gen_duration=0,  # will compute after JSON
                             total_runtime=0)

        # Now capture final timings (after JSON)
        total_end = time.time()
        report_gen_duration = total_end - parsing_end
        total_runtime = total_end - overall_start

        # Update JSON with correct durations (re-write)
        logger.info("Updating JSON with accurate timings...")
        generate_json_report(summary, scanners, config, logger,
                             parsing_duration=parsing_end - overall_start,
                             report_gen_duration=report_gen_duration,
                             total_runtime=total_runtime)

        # Phase 4: Policy evaluation
        exit_code = 0
        if config.fail_on_critical and summary.get_total_critical() > config.max_critical_allowed:
            logger.error(f"❌ Policy failed: {summary.get_total_critical()} critical findings exceed limit")
            exit_code = 1
        if config.fail_on_high and summary.get_total_high() > config.max_high_allowed:
            logger.error(f"❌ Policy failed: {summary.get_total_high()} high findings exceed limit")
            exit_code = 1
        if config.fail_on_any_finding and summary.get_grand_total() > 0:
            logger.error("❌ Policy failed: findings present")
            exit_code = 1

        if exit_code == 0:
            logger.info("✨ All security policies passed")
        else:
            logger.error("Build failed due to security policy violations")

        logger.info(f"Total runtime: {total_runtime:.2f}s")
        return exit_code
    except KeyboardInterrupt:
        logger.warning("Interrupted by user")
        return 130
    except Exception as e:
        logger.error(f"FATAL ERROR: {str(e)}", exc_info=True)
        return 1

if __name__ == "__main__":
    sys.exit(main())
